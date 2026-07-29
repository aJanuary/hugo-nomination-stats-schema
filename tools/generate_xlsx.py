#!/usr/bin/env python3
# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl",
# ]
# ///
"""
Generate an XLSX nomination-statistics workbook, one sheet per category.
"""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# scaledPoints is scaledPoints * 60 to avoid rounding errors under EPH (60 is
# the lowest common denominator of the fractional points possible, since a
# ballot splits one point evenly among up to 5 nominated works).
SCALED_POINTS_DENOMINATOR = 60

# Excel's built-in "Bad" cell style colors, used for elimination-candidate cells.
ELIMINATION_FILL = PatternFill("solid", fgColor="FFC7CE")
ELIMINATION_FONT = Font(color="9C0006")
REMOVED_FONT = Font(strike=True, color="808080")
HEADER_FONT = Font(bold=True)
FOOTNOTE_FONT = Font(italic=True, size=9)
LAST_FINALIST_BORDER = Border(bottom=Side(style="medium", color="000000"))
RIGHT_ALIGN = Alignment(horizontal="right")
LEFT_ALIGN = Alignment(horizontal="left")
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showRowStripes=True,
    showFirstColumn=False,
    showLastColumn=False,
    showColumnStripes=False,
)

INVALID_SHEET_NAME_CHARS = re.compile(r"[\[\]:*?/\\]")
INVALID_TABLE_NAME_CHARS = re.compile(r"[^A-Za-z0-9_]")
MAX_SHEET_NAME_LENGTH = 31


def safe_sheet_name(name, used_names):
    sanitized = INVALID_SHEET_NAME_CHARS.sub(" ", name).strip()
    truncated = sanitized[:MAX_SHEET_NAME_LENGTH]

    candidate = truncated
    suffix = 2
    while candidate in used_names:
        suffix_str = f" ({suffix})"
        candidate = truncated[: MAX_SHEET_NAME_LENGTH - len(suffix_str)] + suffix_str
        suffix += 1

    used_names.add(candidate)
    return candidate


def safe_table_name(sheet_index, sheet_name):
    sanitized = INVALID_TABLE_NAME_CHARS.sub("_", sheet_name)
    if not sanitized or not sanitized[0].isalpha():
        sanitized = f"T{sanitized}"
    return f"{sanitized}_{sheet_index}"


def add_table(sheet, table_name, min_row, max_row, min_col, max_col):
    ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    sheet.add_table(table)


def set_column_widths(sheet, name_field_count, round_count):
    for i in range(name_field_count):
        column = get_column_letter(i + 1)
        sheet.column_dimensions[column].width = 40 if i == 0 else 30

    n_column = get_column_letter(name_field_count + 1)
    sheet.column_dimensions[n_column].width = 6

    for i in range(round_count):
        column = get_column_letter(name_field_count + 2 + i)
        sheet.column_dimensions[column].width = 10


def apply_row_style(sheet, row, min_col, max_col, row_class):
    for column in range(min_col, max_col + 1):
        cell = sheet.cell(row=row, column=column)
        if "finalist" in row_class:
            cell.font = Font(bold=True, strike="removed" in row_class, color=cell.font.color)
        elif "removed" in row_class:
            cell.font = REMOVED_FONT
        if "last-finalist" in row_class:
            cell.border = LAST_FINALIST_BORDER


def render_category_sheet(sheet, category, sheet_index):
    name_fields = category["nameFields"]
    nominees = category["nominees"]
    rounds = category["rounds"]

    footnotes = []

    header_row = 1
    for i, field in enumerate(name_fields):
        cell = sheet.cell(row=header_row, column=i + 1, value=field)
        cell.font = HEADER_FONT
        cell.alignment = LEFT_ALIGN

    n_col = len(name_fields) + 1
    n_cell = sheet.cell(row=header_row, column=n_col, value="N")
    n_cell.font = HEADER_FONT
    n_cell.alignment = RIGHT_ALIGN

    for i, round_ in enumerate(rounds):
        cell = sheet.cell(row=header_row, column=n_col + 1 + i, value=f"R{round_['number']}")
        cell.font = HEADER_FONT
        cell.alignment = RIGHT_ALIGN

    last_finalist_index = max(
        (index for index, nominee in enumerate(nominees) if nominee.get("finalist")),
        default=None,
    )

    rounds_by_index = [
        {entry["nomineeIndex"]: entry for entry in round_["activeNominees"]}
        for round_ in rounds
    ]

    for index, nominee in enumerate(nominees):
        row = header_row + 1 + index

        marker = ""
        if nominee.get("removed") and nominee.get("removedReason"):
            footnotes.append(nominee["removedReason"])
            marker = f" [{len(footnotes)}]"

        for i, _field in enumerate(name_fields):
            value = nominee["name"][i]
            if i == 0:
                value = f"{value}{marker}"
            cell = sheet.cell(row=row, column=i + 1, value=value)
            cell.alignment = LEFT_ALIGN

        n_cell = sheet.cell(row=row, column=n_col, value=nominee["numNominations"])
        n_cell.alignment = RIGHT_ALIGN

        for i, active_by_index in enumerate(rounds_by_index):
            entry = active_by_index.get(index)
            column = n_col + 1 + i
            if entry is None:
                continue
            cell = sheet.cell(
                row=row,
                column=column,
                value=f"={entry['scaledPoints']}/{SCALED_POINTS_DENOMINATOR}",
            )
            cell.number_format = "0.00"
            cell.alignment = RIGHT_ALIGN
            if "status" in entry:
                cell.fill = ELIMINATION_FILL
                cell.font = ELIMINATION_FONT

        row_classes = []
        if nominee.get("finalist"):
            row_classes.append("finalist")
        if nominee.get("removed"):
            row_classes.append("removed")
        if index == last_finalist_index:
            row_classes.append("last-finalist")
        if row_classes:
            apply_row_style(sheet, row, 1, n_col + len(rounds), row_classes)

    last_data_row = header_row + len(nominees)
    add_table(
        sheet,
        safe_table_name(sheet_index, sheet.title),
        header_row,
        last_data_row,
        1,
        n_col + len(rounds),
    )

    next_row = last_data_row + 2
    if footnotes:
        for i, reason in enumerate(footnotes):
            cell = sheet.cell(row=next_row + i, column=1, value=f"{i + 1}. {reason}")
            cell.font = FOOTNOTE_FONT

    set_column_widths(sheet, len(name_fields), len(rounds))
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def render_summary_sheet(sheet, data):
    sheet["A1"] = f"{data['year']} Hugo Award Nomination Statistics"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"{data['numBallots']} ballots cast."

    header_row = 4
    headers = ["Category", "Ballots cast", "Nominees"]
    for i, header in enumerate(headers):
        cell = sheet.cell(row=header_row, column=i + 1, value=header)
        cell.font = HEADER_FONT
        cell.alignment = LEFT_ALIGN if i == 0 else RIGHT_ALIGN

    for index, category in enumerate(data["categories"]):
        row = header_row + 1 + index
        sheet.cell(row=row, column=1, value=category["name"]).alignment = LEFT_ALIGN
        sheet.cell(row=row, column=2, value=category["numBallots"]).alignment = RIGHT_ALIGN
        sheet.cell(row=row, column=3, value=category["numNominees"]).alignment = RIGHT_ALIGN

    last_data_row = header_row + len(data["categories"])
    add_table(sheet, "Summary", header_row, last_data_row, 1, 3)

    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 12
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def render_workbook(data):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    render_summary_sheet(summary_sheet, data)

    used_names = {"Summary"}
    for sheet_index, category in enumerate(data["categories"], start=1):
        sheet_name = safe_sheet_name(category["name"], used_names)
        sheet: Worksheet = workbook.create_sheet(sheet_name)
        render_category_sheet(sheet, category, sheet_index)

    return workbook


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Path to a nominations JSON file")
    parser.add_argument("--output", required=True, help="Path to write the XLSX file")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    data = json.loads(input_path.read_text())
    workbook = render_workbook(data)
    workbook.save(args.output)


if __name__ == "__main__":
    main()
